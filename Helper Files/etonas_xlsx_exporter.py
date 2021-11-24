from parser_utils import get_origin_country, get_total_price, get_sales_channel_hs_code
from parser_constants import ETONAS_HEADERS, ETONAS_HEADERS_MAPPING
import logging
import openpyxl
import sys

# GLOBAL VARIABLES
ETONAS_CHARLIMIT_PER_CELL = 32
VBA_ERROR_ALERT = 'ERROR_CALL_DADDY'
VBA_ETONAS_CHARTLIMIT_ALERT = 'ETONAS_CHARLIMIT_WARNING'


class EtonasExporter():
    '''creates workbook based on Etonas shippment company xlsx requirements
    
    Args:
    -etonas_orders: list of etonas orders (dicts)
    -etonas_path: workbook path to be saved at
    -sales_channel: str option AmazonEU / AmazonCOM / Etsy
    -proxy_keys: dict to handle both Amazon and Etsy sales channels'''

    def __init__(self, etonas_orders : list, etonas_path : str, sales_channel : str, proxy_keys : dict):
        self.etonas_orders = etonas_orders
        self.etonas_path = etonas_path
        self.sales_channel = sales_channel
        self.proxy_keys = proxy_keys
    
    def reformat_data_to_etonas_output(self):
        '''reduces input data to that needed in output csv'''
        try:
            etonas_ready_data = []
            for order_dict in self.etonas_orders:
                reduced_order_dict = self.prepare_etonas_order_contents(order_dict)
                etonas_ready_data.append(reduced_order_dict)
            return etonas_ready_data
        except Exception as e:
            print(VBA_ERROR_ALERT)
            logging.warning(f'Error while iterating collected row dicts and trying to reduce in ETONAS. Error: {e}')

    def prepare_etonas_order_contents(self, order : dict):
        '''returns ready-to-write order data dict based on Etonas file headers'''
        export = {}
        first_name, last_name = self._get_fname_lname(order)
        product_name_proxy_key = self.proxy_keys.get('product-name', '')
        # Change GB to UK for Etonas
        if order[self.proxy_keys['ship-country']] == 'GB':
            order[self.proxy_keys['ship-country']] = 'UK'
        
        for header in ETONAS_HEADERS:
            if header in ETONAS_HEADERS_MAPPING.keys():
                # Etsy has no phone, email, returning empty string, to prevent KeyError
                target_key = self.proxy_keys.get(ETONAS_HEADERS_MAPPING[header], '')
                export[header] = order.get(target_key, '')

                # warn in VBA if char limit per cell is exceeded in Etonas address lines 1/2/3/4
                if 'address' in header.lower() and len(export[header]) > ETONAS_CHARLIMIT_PER_CELL:
                    logging.info(f'Order with key {header} and value {export[header]} triggered VBA warning for charlimit set by Etonas')
                    print(VBA_ETONAS_CHARTLIMIT_ALERT)
            
            elif header == 'First_name':
                export[header] = first_name
            elif header == 'Last_name':
                export[header] = last_name
            elif header == 'HS':
                export[header] = get_sales_channel_hs_code(order, product_name_proxy_key)
            elif header == 'Origin':
                # etsy - no item title, hardcoding for etsy until weight mapping
                if product_name_proxy_key == '':
                    export[header] = 'CN'
                else:
                    export[header] = get_origin_country(order[product_name_proxy_key])
            elif header == 'Currency':
                target_key = self.proxy_keys['currency']
                export[header] = order[target_key].lower()
            elif header == 'Price per quantity':
                export[header] = get_total_price(order, self.sales_channel)
            elif header == 'Weight(Kg)':
                export[header] = self.__get_weight_in_kg(order)
            else:
                export[header] = ''
        return export
    
    def _get_fname_lname(self, order : dict):
        '''returns first and last name based on sales channel'''
        try:
            if self.sales_channel == 'Etsy':
                f_name = order[self.proxy_keys['buyer-fname']]
                l_name = order[self.proxy_keys['buyer-lname']]
                return f_name, l_name
            else:
                f_name, l_name = order[self.proxy_keys['recipient-name']].split(' ', 1)
                return f_name, l_name
        except KeyError as e:
            logging.critical(f'No recipient-name key for etonas func: _get_fname_lname. Err: {e} Order: {order}')
            print(VBA_ERROR_ALERT)
            sys.exit()
        except ValueError as e:
            logging.debug(f'Failed to unpack f_name, l_name for sales ch: {self.sales_channel} etonas xlsx. Err: {e}. Returning proxy recipient-name order val: {order[self.proxy_keys["recipient-name"]]} and empty l_name')
            return order[self.proxy_keys['recipient-name']], ''

    def __get_weight_in_kg(self, order:dict):
        '''returns order weight in kg if possible, empty str if not'''
        try:
            return round(order['weight'] / 1000, 2)
        except:
            return ''

    @staticmethod
    def _write_headers(worksheet, headers):
        for col, header in enumerate(headers, 1):
            worksheet.cell(1, col).value = header    

    @staticmethod
    def range_generator(orders_data, headers):
        for row, _ in enumerate(orders_data):
            for col, _ in enumerate(headers):
                yield row, col
    
    def _write_etonas_orders(self, worksheet, headers, orders_data):
        for row, col in self.range_generator(orders_data, headers):
            working_dict = orders_data[row]
            key_pointer = headers[col]
            # offsets due to excel vs python numbering  + headers in row 1
            worksheet.cell(row + 2, col + 1).value = working_dict[key_pointer]

    def adjust_col_widths(self, worksheet):
        for col in worksheet.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.1
            worksheet.column_dimensions[col_letter].width = adjusted_width

    def export(self):
        reheaded_etonas_orders = self.reformat_data_to_etonas_output()
        wb = openpyxl.Workbook()
        ws = wb.active
        self._write_headers(ws, ETONAS_HEADERS)
        self._write_etonas_orders(ws, ETONAS_HEADERS, reheaded_etonas_orders)        
        self.adjust_col_widths(ws)
        wb.save(self.etonas_path)


if __name__ == "__main__":
    pass