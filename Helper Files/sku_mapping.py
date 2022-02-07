import openpyxl
import logging
import os
from parser_utils import alert_VBA_duplicate_mapping_sku
from excel_utils import get_last_used_row_col
from file_utils import get_output_dir


class ReadExcelFile():
    '''Interface to read excel file, check integrity and return values as dict.
    
    METHODS:
        get_ws_data - returns dict of worksheet A col as keys and B col as values
    
    ARGS:
        config:dict - configuration to read Excel for specific file'''

    def __init__(self, config:dict):
        self.wb_name = config['wb_name']
        self.ws_name = config['ws_name']
        self.start_row = config['start_row']
        self.check_integrity = config['check_integrity']
        self.alert_for_duplicates = config['alert_for_duplicates']
        self.wb_path = os.path.join(get_output_dir(client_file=False), self.wb_name)


    def get_ws_data(self) -> dict:
            '''returns dict of passed config wb/ws values as dict keys for A col and values for B col'''
            try:
                wb = openpyxl.load_workbook(self.wb_path)
                self.ws = wb[self.ws_name]
                self._get_ws_limits()
                if self.check_integrity:
                    self._check_ws_integrity()
                ws_data = self._read_ws_to_dict()
                wb.close()
                return ws_data
            except Exception as e:
                logging.critical(f'Failed to read excel wb: {self.wb_name}. Err: {e}. Closing wb; returning empty dict')
                wb.close()
                return {}
        
    def _get_ws_limits(self):
        '''sets variables self.last_row and self.last_col'''
        ws_limits = get_last_used_row_col(self.ws)
        self.last_col = ws_limits['max_col']
        self.last_row = ws_limits['max_row']

    def _check_ws_integrity(self):
        '''ensures mapping workbook was not structuraly tampered with:
        3 columns, ws name, minimum 50 used rows, header titles'''        
        assert self.last_row > 30, f'Less than 30 rows in SKU Mapping file. Last row used in \'Mapping\' ws: {self.last_row}'
        assert self.last_col == 3, f'Unexpected number of used columns in SKU Mapping file. Expected 3, got {self.last_col}'
        
        a1value = self.ws['A1'].value
        b1value = self.ws['B1'].value
        c1value = self.ws['C1'].value
        assert a1value == 'Amazon SKU', f'Unexpected value {a1value} in SKU Mapping active sheet A1 cell. Expected: Amazon SKU'
        assert b1value == 'Shop4Top Custom Label', f'Unexpected value {b1value} in SKU Mapping active sheet A1 cell. Expected: Shop4Top Custom Label'
        assert c1value == 'Item Title', f'Unexpected value {c1value} in SKU Mapping active sheet A1 cell. Expected: Item Title'

    def _read_ws_to_dict(self) -> dict:
        '''iterates though data rows [<self.start_row>:self.last_row] in self.ws and returns ws_data dict:
        
        In case of SKU_MAPPING WB config: {sku1:custom_label1, sku2:custom_label2, ...}

        In case of SKU_BRAND WB config: {sku1:brand, sku2:brand, ...}
        '''
        ws_data = {}
        for r in range(self.start_row, self.last_row + 1):
            # variable names fit SKU_MAPPING config. For SKU_BRAND sku -> sku, custom_label -> brand
            sku, custom_label = self._get_mapping_row_data(r)            
            if sku not in ws_data.keys():
                ws_data[sku] = custom_label
            else:
                if self.alert_for_duplicates:
                    alert_VBA_duplicate_mapping_sku(sku)
        logging.info(f'Successfuly read {self.wb_name}. Returning dict with {len(ws_data.keys())} entries')
        return ws_data

    def _get_mapping_row_data(self, r:int):
        '''returns two values from columns A,B in self.ws on r (arg) row'''
        col_A_val = self.ws.cell(r, 1).value
        col_B_val = self.ws.cell(r, 2).value
        return col_A_val, col_B_val


if __name__ == "__main__":
    pass