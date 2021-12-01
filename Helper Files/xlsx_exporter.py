from parser_utils import get_origin_country, get_total_price, get_sales_channel_hs_code
from parser_constants import NLPOST_HEADERS, NLPOST_HEADERS_MAPPING, NLPOST_FIXED_VALUES
from parser_constants import ETONAS_HEADERS, ETONAS_HEADERS_MAPPING
from parser_constants import DPDUPS_HEADERS, DPDUPS_HEADERS_MAPPING
import logging
import openpyxl
import sys

# GLOBAL VARIABLES
ETONAS_CHARLIMIT_PER_CELL = 32
NLPOST_CHARLIMIT_PER_CELL = 90
VBA_ERROR_ALERT = 'ERROR_CALL_DADDY'
VBA_ETONAS_CHARTLIMIT_ALERT = 'ETONAS_CHARLIMIT_WARNING'
VBA_NLPOST_CHARTLIMIT_ALERT = 'NLPOST_CHARLIMIT_WARNING'
VBA_MISSING_WEIGHT_DATA_ALERT = 'ETONAS/NLPOST MISSING_WEIGHT_WARNING'
HEADER_SETTINGS = {'etonas': {'headers' : ETONAS_HEADERS, 'mapping': ETONAS_HEADERS_MAPPING},
                'dpdups' : {'headers' : DPDUPS_HEADERS, 'mapping': DPDUPS_HEADERS_MAPPING},
                'nlpost': {'headers' : NLPOST_HEADERS, 'mapping': NLPOST_HEADERS_MAPPING, 'fixed': NLPOST_FIXED_VALUES}}
PACKAGE_DIMENSIONS = {'DKS': {'X': '20', 'Y': '15', 'Z': '10'},
                    'MKS': {'X': '15', 'Y': '10', 'Z': '2'}}


class XlsxExporter():
    '''generic class for creating workbook based on Etonas/NLPost shippment companies xlsx requirements.
    Assumes class that inherit this class have appropriate names as part of class name: (Etonas* / NLPost*)
    
    Args:
    -input_orders: list of orders (dicts) as accepted by class
    -export_path: workbook path to be saved at
    -sales_channel: str option AmazonEU / AmazonCOM / Etsy
    -proxy_keys: dict to handle both Amazon and Etsy sales channels'''

    def __init__(self, input_orders:list, export_path:str, sales_channel:str, proxy_keys:dict):
        self.input_orders = input_orders
        self.export_path = export_path
        self.sales_channel = sales_channel
        self.proxy_keys = proxy_keys
        self.__get_mode()
        self.header_settings = HEADER_SETTINGS[self.mode]
        self.row_offset = 1 if self.mode == 'nlpost' else 0

    def __get_mode(self):
        '''sets self.mode variable to differentiate Etonas / NLPost workbook generation'''
        if self.__class__.__name__.startswith('Etonas'):
            self.mode = 'etonas'
        elif self.__class__.__name__.startswith('NLPost'):
            self.mode = 'nlpost'
        elif self.__class__.__name__.startswith('DPDUPS'):
            self.mode = 'dpdups'
        logging.debug(f'Using XlsxExporter in {self.mode} mode. Parsing {len(self.input_orders)} orders, exporting to path: {self.export_path}')


    def refactor_data_for_export(self):
        '''reduces input data to that needed in output xlsx'''
        try:
            export_ready_data = []
            for order in self.input_orders:
                reduced_order = self._refactor_order(order)
                export_ready_data.append(reduced_order)
            return export_ready_data
        except Exception as e:
            print(VBA_ERROR_ALERT)
            logging.warning(f'Error while iterating collected row dicts and trying to reduce in XlsxExporter mode: {self.mode}. Error: {e}')

    def _refactor_order(self, order:dict) -> dict:
        '''refactors order based on self.mode via prepare_etonas_order_contents or prepare_nlpost_order_contents methods'''
        if self.mode == 'etonas':
            reduced_order = self.prepare_etonas_order_contents(order)
        elif self.mode == 'nlpost':
            reduced_order = self.prepare_nlpost_order_contents(order)
        elif self.mode == 'dpdups':
            reduced_order = self.prepare_dpdups_order_contents(order)
        return reduced_order

    def prepare_etonas_order_contents(self, order:dict) -> dict:
        '''implemented in inheriting class'''
        logging.warning(f'You should not be using generic class to create xlsx output. Warning from: prepare_etonas_order_contents method')
        return order
    
    def prepare_nlpost_order_contents(self, order:dict) -> dict:
        '''implemented in inheriting class'''
        logging.warning(f'You should not be using generic class to create xlsx output. Warning from: prepare_nlpost_order_contents method')
        return order

    def prepare_dpdups_order_contents(self, order:dict) -> dict:
        '''implemented in inheriting class'''
        logging.warning(f'You should not be using generic class to create xlsx output. Warning from: prepare_nlpost_order_contents method')
        return order

    def _get_fname_lname(self, order:dict):
        '''returns first and last name based on sales channel'''
        try:
            if self.sales_channel == 'Etsy':
                f_name = order[self.proxy_keys['buyer-fname']]
                l_name = order[self.proxy_keys['buyer-lname']]
                return f_name, l_name
            else:
                f_name, l_name = order[self.proxy_keys['recipient-name']].split(' ', 1)
                return f_name, l_name
        except KeyError as e:
            logging.critical(f'No recipient-name key for etonas func: _get_fname_lname. Err: {e} Order: {order}')
            print(VBA_ERROR_ALERT)
            sys.exit()
        except ValueError as e:
            logging.debug(f'Failed to unpack f_name, l_name for sales ch: {self.sales_channel} etonas xlsx. Err: {e}. Returning proxy recipient-name order val: {order[self.proxy_keys["recipient-name"]]} and empty l_name')
            return order[self.proxy_keys['recipient-name']], ''

    def _get_weight_in_kg(self, order:dict):
        '''returns order weight in kg if possible, empty str if not'''
        try:
            return round(order['weight'] / 1000, 3)
        except:
            print(VBA_MISSING_WEIGHT_DATA_ALERT)
            return ''

    def _write_headers(self, ws:object, headers:list):
        for col, header in enumerate(headers, 1):
            ws.cell(1 + self.row_offset, col).value = header    

    @staticmethod
    def range_generator(orders:list, headers:list):
        for row, _ in enumerate(orders):
            for col, _ in enumerate(headers):
                yield row, col
    
    def _write_orders(self, ws:object, headers:list, orders:list):
        for row, col in self.range_generator(orders, headers):
            working_dict = orders[row]
            key_pointer = headers[col]
            # offsets due to excel vs python numbering  + headers in row 1 + self.row_offset (first empty row for nlpost)
            ws.cell(row + 2 + self.row_offset, col + 1).value = working_dict[key_pointer]

    def adjust_col_widths(self, ws:object):
        '''iterates cols, cells within col, adjusts column width based on max char cell within col + extra spacing'''
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.1
            ws.column_dimensions[col_letter].width = adjusted_width

    def export(self):
        export_data = self.refactor_data_for_export()
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = self.header_settings['headers']
        self._write_headers(ws, headers)
        self._write_orders(ws, headers, export_data)
        self.adjust_col_widths(ws)
        wb.save(self.export_path)
        wb.close()



class NLPostExporter(XlsxExporter):
    '''Creates Excel orders workbook based on NLPost xlsx requirements. Uses generic parent XlsxExporter class. 
    class name must include word 'NLPost' (not actually, but kepping it clean).
    
    only overwritten method: prepare_nlpost_order_contents
    
    Args:
    -input_orders: list of orders (dicts) as accepted by class
    -export_path: workbook path to be saved at
    -sales_channel: str option AmazonEU / AmazonCOM / Etsy
    -proxy_keys: dict to handle both Amazon and Etsy sales channels'''


    def prepare_nlpost_order_contents(self, order:dict) -> dict:
        '''returns ready-to-write order data dict based on NLPost file headers'''
        export = {}
        for header in NLPOST_HEADERS:
            if header in NLPOST_FIXED_VALUES.keys():
                export[header] = NLPOST_FIXED_VALUES[header]
            elif header in NLPOST_HEADERS_MAPPING.keys():
                # Etsy has no phone, email, returning empty string, to prevent KeyError
                target_key = self.proxy_keys.get(NLPOST_HEADERS_MAPPING[header], '')
                raw_target_value = order.get(target_key, '')

                if header == 'Receiver postal code' or header == 'Receiver phone':
                    # strip special chars from postal code and phone number
                    export[header] = self.__strip_special_nlpost_chars(raw_target_value)
                elif header == 'Description':
                    export[header] = self.__make_batteries_alkaline(raw_target_value)
                else:
                    export[header] = raw_target_value

            elif header == 'Receiver street':
                # combine two (three for amazon) address fields
                address1 = order[self.proxy_keys['ship-address-1']]
                address2 = order[self.proxy_keys['ship-address-2']]
                address = f'{address1} {address2} {order[self.proxy_keys["ship-address-3"]]}' if self.sales_channel != 'Etsy' else f'{address1} {address2}'
                export[header] = address

                if len(address) > NLPOST_CHARLIMIT_PER_CELL:
                    logging.warning(f'Order with key {header} and value {address} triggered VBA warning for charlimit set by NLPost')
                    print(VBA_NLPOST_CHARTLIMIT_ALERT)
            elif header in ['X', 'Y', 'Z']:
                parcel_dimension = self.__get_package_dimension(order['vmdoption'], header)
                export[header] = parcel_dimension
            elif header == 'Service name':
                export[header] = self.__get_nlpost_service_name(order)
            elif header == 'Weight':
                export[header] = self._get_weight_in_kg(order)    
            elif header == 'HS code':
                product_name_proxy_key = self.proxy_keys.get('title', '')
                export[header] = get_sales_channel_hs_code(order, product_name_proxy_key)
            else:
                export[header] = ''
        return export

    def __get_package_dimension(self, vmdoption:str, header:str) -> str:
        '''returns package dimension in cm, formatted for NLPost'''
        if vmdoption not in ['VKS', 'MKS', 'DKS']:
            print(VBA_MISSING_WEIGHT_DATA_ALERT)
            return ''
        package_category = 'DKS' if vmdoption == 'DKS' else 'MKS'
        return PACKAGE_DIMENSIONS[package_category][header]

    def __strip_special_nlpost_chars(self, text:str) -> str:
        '''removes characters from 'text' arg not allowed in nlpost system'''
        return text.replace('(', '').replace(')', '').replace('+', '').replace(';', '').replace(':', '').replace(' ', '').replace('-', '')

    def __get_nlpost_service_name(self, order:dict) -> str:
        '''returns value for NLPost column 'Service name' based on order details'''
        if order['vmdoption'] == '':
            return 'No Data'

        if order['tracked']:
            return 'PS4'
        else:
            # PEC0 for all Untracked orders
            return 'PEC0'

    def __make_batteries_alkaline(self, contents:str) -> str:
        '''replaces batteries with alkaline batteries string if applicable'''
        return contents.replace('BATTERIES', 'ALKALINE BATTERIES')



class EtonasExporter(XlsxExporter):
    '''Creates Excel orders workbook based on Etonas xlsx requirements. Uses generic parent XlsxExporter class. 
    class name must include word 'Etonas'.
    
    only overwritten method: prepare_etonas_order_contents
    
    Args:
    -input_orders: list of orders (dicts) as accepted by class
    -export_path: workbook path to be saved at
    -sales_channel: str option AmazonEU / AmazonCOM / Etsy
    -proxy_keys: dict to handle both Amazon and Etsy sales channels'''


    def prepare_etonas_order_contents(self, order:dict) -> dict:
        '''returns ready-to-write order data dict based on Etonas file headers'''
        export = {}
        first_name, last_name = self._get_fname_lname(order)
        product_name_proxy_key = self.proxy_keys.get('title', '')

        # Change GB to UK for Etonas
        if order[self.proxy_keys['ship-country']] == 'GB':
            order[self.proxy_keys['ship-country']] = 'UK'
        
        for header in ETONAS_HEADERS:
            if header in ETONAS_HEADERS_MAPPING.keys():
                # Etsy has no phone, email, returning empty string, to prevent KeyError
                target_key = self.proxy_keys.get(ETONAS_HEADERS_MAPPING[header], '')
                export[header] = order.get(target_key, '')

                # warn in VBA if char limit per cell is exceeded in Etonas address lines 1/2/3/4
                if 'address' in header.lower() and len(export[header]) > ETONAS_CHARLIMIT_PER_CELL:
                    logging.warning(f'Order with key {header} and value {export[header]} triggered VBA warning for charlimit set by Etonas')
                    print(VBA_ETONAS_CHARTLIMIT_ALERT)
            
            elif header == 'First_name':
                export[header] = first_name
            elif header == 'Last_name':
                export[header] = last_name
            elif header == 'GLS':
                export[header] = '1' if order[self.proxy_keys['ship-country']] in ['UK', 'DE'] else '0'
            elif header == 'Tracking (0 - neregistruota, 1 - registruota)':
                export[header] = int(order['tracked'])
            elif header == 'HS':
                export[header] = get_sales_channel_hs_code(order, product_name_proxy_key)
            elif header == 'Origin':
                if product_name_proxy_key == '':
                    export[header] = 'CN'
                else:
                    export[header] = get_origin_country(order[product_name_proxy_key])
            elif header == 'Currency':
                target_key = self.proxy_keys['currency']
                export[header] = order[target_key].lower()
            elif header == 'Price per quantity':
                export[header] = get_total_price(order, self.sales_channel)
            elif header == 'Weight(Kg)':
                export[header] = self._get_weight_in_kg(order)
            else:
                export[header] = ''
        return export


class DPDUPSExporter(XlsxExporter):
    '''Creates Excel orders workbook for DPD / UPS orders. Uses generic parent XlsxExporter class. 
    class name must include word 'DPDUPS'.
    
    only overwritten method: prepare_dpdups_order_contents
    
    Args:
    -input_orders: list of orders (dicts) as accepted by class
    -export_path: workbook path to be saved at
    -sales_channel: str option AmazonEU / AmazonCOM / Etsy
    -proxy_keys: dict to handle both Amazon and Etsy sales channels'''
    
 
    def prepare_dpdups_order_contents(self, order:dict) -> dict:
        '''returns ready-to-write order data dict based on DPDUPS file headers'''
        export = {}        
        for header in DPDUPS_HEADERS:
            if header in DPDUPS_HEADERS_MAPPING.keys():
                target_key = self.proxy_keys.get(DPDUPS_HEADERS_MAPPING[header], '')
                export[header] = order.get(target_key, '')

            elif header == 'Service Picked':
                export[header] = order['shipping_service']
            elif header == 'Tracked':
                export[header] = order['tracked']
            elif header == 'Sales Channel':
                export[header] = self.sales_channel

        return export


if __name__ == "__main__":
    pass