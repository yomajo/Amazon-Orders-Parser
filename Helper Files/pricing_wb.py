from excel_utils import get_last_used_row_col, cell_to_float
from file_utils import get_output_dir
from countries import COUNTRY_CODES
import openpyxl
import logging
import os


# GLOBAL VARIABLES
PRICING_WB = 'PRICING.xlsx'
ALLOWED_SERVICE_QUERIES = ['NL', 'LP', 'DP', 'ETONAS', 'DPD', 'UPS']


class PricingWB:
    '''interaction with PRICING.xlsx workbook. Assumes workbook integrity has been checked on VBA side.
    
    Args:
    proxy_keys:dict (order key mapping for Amazon / Etsy)

    main method:
    get_pricing_offer - returns price offer as float if found, None otherwise'''

    def __init__(self, proxy_keys:dict):
        self.proxy_keys = proxy_keys
        wb_path = os.path.join(get_output_dir(client_file=False), PRICING_WB)
        self.wb = openpyxl.load_workbook(wb_path, data_only=True)
        self.ws_tracked = self.wb['PrTracked']
        self.ws_untracked = self.wb['PrUntracked']
        self.ws_tracked_limits = get_last_used_row_col(self.ws_tracked)
        self.ws_untracked_limits = get_last_used_row_col(self.ws_untracked)

    def get_pricing_offer(self, order:dict, service:str):
        '''returns price offer for order data provided. External error handling, allow to fail here'''
        tracked, country_code = order['tracked'], order[self.proxy_keys['ship-country']]
        logging.debug(f'Getting offer for: {service}. Tracked: {tracked}, country: {country_code}')
        self.__validate_query(service, country_code)
        ws = self.ws_tracked if tracked else self.ws_untracked
        limits = self.ws_tracked_limits if tracked else self.ws_untracked_limits
        target_row = self.__get_country_row(ws, limits, country_code)
        target_col = self.__get_target_col(ws, limits, order, service)
        
        offer = ws.cell(row=target_row, column=target_col).value
        logging.debug(f'returning offer before float conversion: {offer}')
        return cell_to_float(offer)

    def __validate_query(self, service:str, country_code:str):
        '''validates external querying for basic compatibility with pricing sheets'''
        if service not in ALLOWED_SERVICE_QUERIES:
            logging.critical(f'Pricing was queried by unsupported service: {service}')
            raise ValueError('Order pricing: Service not supported')
        if country_code not in COUNTRY_CODES.values():
            logging.warning(f'Attempt to query pricing for not supported country: {country_code}')
            raise ValueError('Order pricing: Country code not supported')

    def __get_country_row(self, ws:object, limits:dict, country_code:str) -> int:
        '''returns country matching row inside ws sheet. 0 if not found'''
        max_row = limits['max_row']
        for row in range(1, max_row + 1):
            if ws.cell(row=row, column=1).value == country_code:                
                return row
        return 0

    def __get_target_col(self, ws:object, limits:dict, order:dict, service:str) -> int:
        '''returns target column for service based on order and pricing sheet data'''
        max_col = limits['max_col']
        order_weight, vmdoption = order['weight'], order['vmdoption']
        logging.debug(f'Weight: {order_weight}, vmdoption: {vmdoption}')
        segment_start_col = self.__get_segment_start_col(ws, max_col, service)
        segment_end_col = self.__get_segment_end_col(ws, segment_start_col)
        vmdoption = self.__validate_vmdoption(ws, vmdoption, segment_start_col, segment_end_col)
        adj_start_col = self.__get_vmd_adj_start_col(ws, vmdoption, segment_start_col, segment_end_col)
        
        # find target column
        for col in range(adj_start_col, segment_end_col + 1):
            col_weight_limit = ws.cell(row=3, column=col).value
            if order_weight <= col_weight_limit:
                return col
        return 0

    def __get_segment_start_col(self, ws:object, max_col:int, service:str) -> int:
        '''returns segment start column for target row and service. 0 if not found'''
        for col in range(2, max_col + 1):
            if ws.cell(row=1, column=col).value == service:
                return col
        return 0

    def __get_segment_end_col(self, ws:object, segment_start_col:int) -> int:
        '''returns last column in service segment (search range end'''
        for col in range(segment_start_col, segment_start_col + 50):
            if ws.cell(row=2, column=col).value == None:
                return col - 1
        return 0

    def __validate_vmdoption(self, ws:object, vmdoption:str, segment_start_col:int, segment_end_col:int) -> str:
        '''if Shipping service does not support VKS / MKS return next available by hierarchy: VKS -> MKS -> DKS
        returns first vmdtoption that is not less than original and available within service segment columns'''
        # least bad (?) way to introduce and upgrade string hierarchy with indexing
        logging.debug(f'Entering VMD validation with vmdoption arg value: {vmdoption}')
        vmd_hierarchy = {'VKS': 1, 'MKS': 2, 'DKS': 3}
        vmd_options = {1: 'VKS', 2: 'MKS', 3: 'DKS'}
        order_vmd_idx = vmd_hierarchy[vmdoption]
        for col in range(segment_start_col, segment_end_col + 1):
            if ws.cell(row=2, column=col).value == vmdoption:
                return vmdoption

        # if not found, increase index, recursively call with upgraded vmdoption
        upgraded_vmd = vmd_options[order_vmd_idx + 1]
        logging.debug(f'No {vmdoption} match found, upgrading to {upgraded_vmd}. CALLING RECCURSIVELY')
        return self.__validate_vmdoption(ws, upgraded_vmd, segment_start_col, segment_end_col)

    def __get_vmd_adj_start_col(self, ws:object, vmdoption:str, segment_start_col:int, segment_end_col:int) -> int:
        '''returns service segment start col as int, adjusted for vmdoption actually available inside segment headers'''
        for col in range(segment_start_col, segment_end_col + 1):
            if ws.cell(row=2, column=col).value == vmdoption:
                return col
        return 0


if __name__ == '__main__':
    pass